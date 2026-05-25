from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.cleaners import clean_document, clean_email, normalize_column_name
from src.config import DEFAULT_SHEET, EXPECTED_COLUMNS, INPUT_PATH, MAX_EMAILS_PER_DOCUMENT, OUTPUT_PATH
from src.exceptions import (
    EmptyExcelError,
    ExcelReadError,
    ExcelWriteError,
    InputFileNotFoundError,
    MissingColumnsError,
)
from src.risk_rules import evaluate_email


Logger = Callable[[str], None]


@dataclass
class ProcessingResult:
    output_path: Path
    total_records_read: int
    total_records_processed: int
    total_valid_emails: int
    warnings: list[str]


def _log(logger: Logger | None, message: str) -> None:
    if logger:
        logger(message)


def read_input_excel(
    input_path: Path = INPUT_PATH,
    sheet_name: int | str = DEFAULT_SHEET,
    logger: Logger | None = None,
) -> tuple[pd.DataFrame, int]:
    if not input_path.exists():
        raise InputFileNotFoundError(f"No se encontro el archivo de entrada: {input_path}")

    try:
        dataframe = pd.read_excel(input_path, sheet_name=sheet_name, dtype=str)
    except Exception as error:
        raise ExcelReadError(f"No se pudo leer el archivo Excel: {input_path}") from error

    if dataframe.empty:
        raise EmptyExcelError("El Excel de entrada esta vacio.")

    total_records_read = len(dataframe)
    dataframe.columns = [normalize_column_name(column) for column in dataframe.columns]

    _log(logger, "Validando columnas requeridas.")
    missing_columns = [column for column in EXPECTED_COLUMNS if column not in dataframe.columns]
    if missing_columns:
        raise MissingColumnsError(
            "Faltan columnas obligatorias: "
            f"{', '.join(missing_columns)}. Columnas encontradas: {list(dataframe.columns)}"
        )

    dataframe = dataframe[list(EXPECTED_COLUMNS)].copy()
    dataframe["DOCUMENTO"] = dataframe["DOCUMENTO"].apply(clean_document)
    dataframe["CORREO"] = dataframe["CORREO"].apply(clean_email)

    processed_dataframe = dataframe[(dataframe["DOCUMENTO"] != "") & (dataframe["CORREO"] != "")]
    if processed_dataframe.empty:
        raise EmptyExcelError("No hay registros procesables con DOCUMENTO y CORREO.")

    return processed_dataframe, total_records_read


def build_validation_dataframe(input_dataframe: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for _, row in input_dataframe.iterrows():
        evaluation = evaluate_email(row["CORREO"])
        rows.append(
            {
                "DOCUMENTO": row["DOCUMENTO"],
                "CORREO_ORIGINAL": row["CORREO"],
                **evaluation,
            }
        )

    return pd.DataFrame(rows)


def build_grouped_valid_emails(
    validation_dataframe: pd.DataFrame,
    max_emails_per_document: int = MAX_EMAILS_PER_DOCUMENT,
) -> pd.DataFrame:
    columns = ["DOCUMENTO", *[f"CORREO {index}" for index in range(1, max_emails_per_document + 1)]]

    if validation_dataframe.empty:
        return pd.DataFrame(columns=columns)

    valid_emails = validation_dataframe[validation_dataframe["VALIDACION"] == "VALIDO"].copy()
    valid_emails = valid_emails.drop_duplicates(subset=["DOCUMENTO", "CORREO_LIMPIO"])

    grouped_rows = []
    for document, group in valid_emails.groupby("DOCUMENTO"):
        emails = group["CORREO_LIMPIO"].tolist()[:max_emails_per_document]
        grouped_rows.append(
            {
                "DOCUMENTO": document,
                **{
                    f"CORREO {index + 1}": emails[index] if index < len(emails) else ""
                    for index in range(max_emails_per_document)
                },
            }
        )

    return pd.DataFrame(grouped_rows, columns=columns)


def build_summary_dataframe(validation_dataframe: pd.DataFrame) -> pd.DataFrame:
    return build_summary_dataframe_with_totals(validation_dataframe, len(validation_dataframe))


def build_summary_dataframe_with_totals(
    validation_dataframe: pd.DataFrame,
    total_records_read: int,
    processed_at: datetime | None = None,
) -> pd.DataFrame:
    processing_time = processed_at or datetime.now()

    if validation_dataframe.empty:
        counters = {
            "Total de registros leidos": total_records_read,
            "Total de registros procesados": 0,
            "Total de correos validos": 0,
            "Total riesgo medio": 0,
            "Total alto riesgo de rebote": 0,
            "Total no validos": 0,
            "Total dominios unicos": 0,
            "Total dominios sin MX": 0,
            "Total dominios temporales": 0,
            "Total dominios posiblemente mal escritos": 0,
        }
    else:
        counters = {
            "Total de registros leidos": total_records_read,
            "Total de registros procesados": len(validation_dataframe),
            "Total de correos validos": (validation_dataframe["VALIDACION"] == "VALIDO").sum(),
            "Total riesgo medio": (validation_dataframe["VALIDACION"] == "RIESGO MEDIO").sum(),
            "Total alto riesgo de rebote": (
                validation_dataframe["VALIDACION"] == "ALTO RIESGO DE REBOTE"
            ).sum(),
            "Total no validos": (validation_dataframe["VALIDACION"] == "NO VALIDO").sum(),
            "Total dominios unicos": validation_dataframe["DOMINIO"].replace("", pd.NA).nunique(),
            "Total dominios sin MX": (
                (validation_dataframe["FORMATO_VALIDO"] == "SI")
                & (validation_dataframe["MX_VALIDO"] == "NO")
            ).sum(),
            "Total dominios temporales": (validation_dataframe["DOMINIO_TEMPORAL"] == "SI").sum(),
            "Total dominios posiblemente mal escritos": (
                validation_dataframe["DOMINIO_SUGERIDO"].fillna("") != ""
            ).sum(),
        }

    rows = [{"INDICADOR": indicator, "VALOR": int(value)} for indicator, value in counters.items()]
    rows.append(
        {
            "INDICADOR": "Fecha y hora de procesamiento",
            "VALOR": processing_time.strftime("%Y-%m-%d %H:%M:%S"),
        }
    )

    return pd.DataFrame(rows)


def format_output_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book

    for worksheet in workbook.worksheets:
        if worksheet.max_row > 0:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        document_column_index = None
        for cell in worksheet[1]:
            if cell.value == "DOCUMENTO":
                document_column_index = cell.column
                break

        if document_column_index:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=document_column_index).number_format = "@"

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def write_output_excel(
    validation_dataframe: pd.DataFrame,
    grouped_dataframe: pd.DataFrame,
    summary_dataframe: pd.DataFrame,
    output_path: Path = OUTPUT_PATH,
) -> None:
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            validation_dataframe.to_excel(writer, sheet_name="Validacion", index=False)
            grouped_dataframe.to_excel(writer, sheet_name="Correos_Agrupados", index=False)
            summary_dataframe.to_excel(writer, sheet_name="Resumen", index=False)
            format_output_workbook(writer)
    except Exception as error:
        raise ExcelWriteError(f"No se pudo escribir el archivo de salida: {output_path}") from error


def process_excel(
    input_path: Path = INPUT_PATH,
    output_path: Path = OUTPUT_PATH,
    max_emails_per_document: int = MAX_EMAILS_PER_DOCUMENT,
    sheet_name: int | str = DEFAULT_SHEET,
    logger: Logger | None = None,
) -> ProcessingResult:
    warnings = []

    _log(logger, "Leyendo archivo de entrada.")
    input_dataframe, total_records_read = read_input_excel(input_path, sheet_name, logger)

    _log(logger, f"Registros leidos: {total_records_read}.")
    _log(logger, f"Registros procesables: {len(input_dataframe)}.")

    _log(logger, "Procesando correos.")
    validation_dataframe = build_validation_dataframe(input_dataframe)

    _log(logger, "Generando hoja de validacion.")
    total_valid_emails = int((validation_dataframe["VALIDACION"] == "VALIDO").sum())
    if total_valid_emails == 0:
        warnings.append("No se encontraron correos validos para agrupar.")

    _log(logger, "Generando hoja de correos agrupados.")
    grouped_dataframe = build_grouped_valid_emails(validation_dataframe, max_emails_per_document)

    _log(logger, "Generando hoja resumen.")
    summary_dataframe = build_summary_dataframe_with_totals(validation_dataframe, total_records_read)

    _log(logger, "Escribiendo archivo de salida.")
    write_output_excel(validation_dataframe, grouped_dataframe, summary_dataframe, output_path)

    _log(logger, "Archivo generado correctamente.")
    return ProcessingResult(
        output_path=output_path,
        total_records_read=total_records_read,
        total_records_processed=len(validation_dataframe),
        total_valid_emails=total_valid_emails,
        warnings=warnings,
    )
